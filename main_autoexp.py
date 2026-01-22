# main_autoexp.py - 자동 실험 (간소화 버전)
# @ 및 ^ 실험: 3가지 best model (AUROC, AUPRC, Last) 저장 후 한번에 테스트

import os
import time
import copy
from collections import Counter
from datetime import datetime

import numpy as np
import pandas as pd
import torch
from torch.utils.data import DataLoader

from utils import set_seed, load_or_extract_data
from model import get_model
from dataloader import ECGDataset
from train import train_one_epoch, validate
from test import evaluate, calculate_metrics
from openpyxl import load_workbook
from logger import (
    TrainingLogger, print_epoch_header, print_per_class_metrics,
    print_epoch_stats, print_confidence_stats, print_epoch_time,
    calculate_auprc, calculate_auroc
)

# =============================================================================
# 실험 설정
# =============================================================================

# 공통 설정
DATA_PATH = './data/mit-bih-arrhythmia-database-1.0.0/'
OUTPUT_PATH = './auto_results_batch1024/'
BATCH_SIZE = 1024
EPOCHS = 50
LR = 0.0001
WEIGHT_DECAY = 1e-3
SEED = 1234
POLY1_EPS = 0.0
POLY2_EPS = 0.0
CLASSES = ['N', 'S', 'V', 'F']

# RR Feature 설정
RR_FEATURE_OPTION = "opt3"
RR_FEATURE_DIMS = {"opt1": 7, "opt2": 38, "opt3": 7, "opt4": 7}

# 모델 설정
MODEL_CONFIG = {
    'in_channels': 1,
    'out_ch': 180,
    'mid_ch': 30,
    'num_heads': 9,
    'n_rr': RR_FEATURE_DIMS[RR_FEATURE_OPTION],
}

# ECG Parameters
VALID_LEADS = ['MLII', 'V1', 'V2', 'V4', 'V5']
OUT_LEN = 720

# 데이터 분할
DS1_TRAIN_SPLIT = [
    '101', '106', '108', '109', '112', '115', '116', '118', '119',
    '122', '201', '203', '209', '215', '223', '230'
]
DS1_VALID_SPLIT = ['114', '124', '205', '207', '220', '208']
DS2_TEST = [
    '100', '103', '105', '111', '113', '117', '121', '123', '200', '202',
    '210', '212', '213', '214', '219', '221', '222', '228', '231', '232',
    '233', '234'
]

# =============================================================================
# 실험 정의
# =============================================================================

# 실험 목록: (실험명, 모델타입)
EXPERIMENTS = [
    # A 시리즈 (Dense Block 포함)
    # ('A0', 'baseline'),
    ('B2', 'cross_attention_B'),
    # ('A1', 'naive_concatenate'),
    # ('A2', 'cross_attention'),

    # # B 시리즈 (Dense Block 없음)
    # ('B0', 'baseline_B'),
    # ('B1', 'naive_concatenate_B'),
    # ('B2', 'cross_attention_B'),
]

# =============================================================================
# 유틸리티 함수
# =============================================================================

def create_auto_exp_dir(exp_name):
    """자동 실험용 폴더 생성"""
    os.makedirs(OUTPUT_PATH, exist_ok=True)
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    exp_dir = os.path.join(OUTPUT_PATH, f'{exp_name}_{timestamp}')
    os.makedirs(exp_dir, exist_ok=True)
    os.makedirs(os.path.join(exp_dir, 'best_weights'), exist_ok=True)
    return exp_dir


def run_experiment(exp_name, model_type, device):
    """
    단일 실험 수행: AUROC/AUPRC/Last 3개 모델 저장 후 한번에 테스트

    Args:
        exp_name: 실험 이름 (A0, A1, B0, 등)
        model_type: 모델 타입 (baseline, naive_concatenate, cross_attention, *_B)
        device: torch device

    Returns:
        results_dict: 3가지 모델별 테스트 결과 딕셔너리
    """
    print(f"\n{'='*80}")
    print(f"Experiment: {exp_name} (Model: {model_type})")
    print(f"{'='*80}")

    set_seed(SEED)
    exp_dir = create_auto_exp_dir(exp_name)

    # TensorBoard Logger
    logger = TrainingLogger(os.path.join(exp_dir, 'runs'))

    # 데이터 로드 (DS1-1 train, DS1-2 valid, DS2 test)
    train_data, train_labels, train_rr, train_pid, train_sid = load_or_extract_data(
        record_list=DS1_TRAIN_SPLIT, base_path=DATA_PATH, valid_leads=VALID_LEADS,
        out_len=OUT_LEN, split_name=f"Train_{exp_name}"
    )
    valid_data, valid_labels, valid_rr, valid_pid, valid_sid = load_or_extract_data(
        record_list=DS1_VALID_SPLIT, base_path=DATA_PATH, valid_leads=VALID_LEADS,
        out_len=OUT_LEN, split_name=f"Valid_{exp_name}"
    )
    test_data, test_labels, test_rr, test_pid, test_sid = load_or_extract_data(
        record_list=DS2_TEST, base_path=DATA_PATH, valid_leads=VALID_LEADS,
        out_len=OUT_LEN, split_name=f"Test_{exp_name}"
    )

    # DataLoader 생성
    train_dataset = ECGDataset(train_data, train_rr, train_labels, train_pid, train_sid)
    valid_dataset = ECGDataset(valid_data, valid_rr, valid_labels, valid_pid, valid_sid)
    test_dataset = ECGDataset(test_data, test_rr, test_labels, test_pid, test_sid)


    train_loader = DataLoader(train_dataset, batch_size=BATCH_SIZE, shuffle=True, 
                              num_workers=4, pin_memory=True)
    valid_loader = DataLoader(valid_dataset, batch_size=BATCH_SIZE, shuffle=False,
                              num_workers=4, pin_memory=True)
    test_loader = DataLoader(test_dataset, batch_size=BATCH_SIZE, shuffle=False,
                             num_workers=4, pin_memory=True)

    print(f"  Train: {len(train_labels):,} samples | {dict(Counter(train_labels))}")
    print(f"  Valid: {len(valid_labels):,} samples | {dict(Counter(valid_labels))}")
    print(f"  Test : {len(test_labels):,} samples | {dict(Counter(test_labels))}")

    n_records = len(DS1_TRAIN_SPLIT) + len(DS1_VALID_SPLIT)

    # 모델 생성
    model = get_model(
        exp_name=model_type,
        nOUT=len(CLASSES),
        n_pid=n_records,
        **MODEL_CONFIG
    ).to(device)

    # 학습 설정
    optimizer = torch.optim.AdamW(model.parameters(), lr=LR, weight_decay=WEIGHT_DECAY)
    scheduler = torch.optim.lr_scheduler.StepLR(optimizer, step_size=20, gamma=0.5)

    # Best 모델 추적 (AUROC, AUPRC, Last)
    best = {
        'auroc': {'value': 0.0, 'epoch': 0, 'state_dict': None, 'valid_metrics': None},
        'auprc': {'value': 0.0, 'epoch': 0, 'state_dict': None, 'valid_metrics': None},
        'last': {'epoch': EPOCHS, 'state_dict': None, 'valid_metrics': None}
    }

    # 학습 루프
    print_epoch_header()

    for epoch in range(1, EPOCHS + 1):
        epoch_start = time.time()

        # Train
        train_loss, train_metrics, *p_t_train = train_one_epoch(
            model, train_loader, POLY1_EPS, POLY2_EPS, optimizer, device
        )
        current_lr = optimizer.param_groups[0]['lr']

        print_epoch_stats(epoch, train_loss, train_metrics['acc'], current_lr, phase='Train')
        print_confidence_stats(*p_t_train, phase='Train')
        logger.log_epoch(epoch, train_loss, train_metrics, phase='train')
        logger.log_confidence(epoch, *p_t_train, phase='train')

        # Validation
        valid_loss, valid_metrics, *p_t_valid = validate(
            model, valid_loader, POLY1_EPS, POLY2_EPS, device
        )

        print_epoch_stats(epoch, valid_loss, valid_metrics['acc'], current_lr, phase='Valid')
        print_confidence_stats(*p_t_valid, phase='Valid')
        logger.log_epoch(epoch, valid_loss, valid_metrics, phase='valid')
        logger.log_confidence(epoch, *p_t_valid, phase='valid')

        # Best AUROC 체크
        if valid_metrics['macro_auroc'] > best['auroc']['value']:
            best['auroc'] = {
                'value': valid_metrics['macro_auroc'],
                'epoch': epoch,
                'state_dict': copy.deepcopy(model.state_dict()),
                'valid_metrics': copy.deepcopy(valid_metrics)
            }
            print(f"  ★ [BEST AUROC] {best['auroc']['value']:.4f}")

        # Best AUPRC 체크
        if valid_metrics['macro_auprc'] > best['auprc']['value']:
            best['auprc'] = {
                'value': valid_metrics['macro_auprc'],
                'epoch': epoch,
                'state_dict': copy.deepcopy(model.state_dict()),
                'valid_metrics': copy.deepcopy(valid_metrics)
            }
            print(f"  ★ [BEST AUPRC] {best['auprc']['value']:.4f}")

        # Last epoch 저장
        if epoch == EPOCHS:
            best['last'] = {
                'epoch': epoch,
                'state_dict': copy.deepcopy(model.state_dict()),
                'valid_metrics': copy.deepcopy(valid_metrics)
            }

        scheduler.step()
        print_epoch_time(epoch, time.time() - epoch_start)
        print("=" * 80)

    logger.close()

    # ========================================================================
    # 학습 완료 후 3가지 모델로 Test 수행 및 TensorBoard 로깅
    # ========================================================================
    print(f"\n{'='*80}")
    print(f"Testing 3 Best Models: AUROC, AUPRC, Last")
    print(f"{'='*80}")

    results_dict = {}

    for model_key in ['auroc', 'auprc', 'last']:
        epoch = best[model_key]['epoch']
        print(f"\n--- Testing {model_key.upper()} Best Model (Epoch {epoch}) ---")

        # 모델 로드
        model.load_state_dict(best[model_key]['state_dict'])
        model.eval()

        # Test 수행
        y_pred, y_true, eval_results = evaluate(model, test_loader, device)
        test_metrics = calculate_metrics(np.array(y_true), np.array(y_pred))

        # AUROC/AUPRC 계산
        y_probs = np.array([item['all_probs'] for item in eval_results])
        num_classes = len(CLASSES)

        macro_auprc, weighted_auprc, per_class_auprc = calculate_auprc(
            np.array(y_true), y_probs, num_classes
        )
        macro_auroc, weighted_auroc, per_class_auroc = calculate_auroc(
            np.array(y_true), y_probs, num_classes
        )

        test_metrics['macro_auroc'] = macro_auroc
        test_metrics['macro_auprc'] = macro_auprc
        test_metrics['weighted_auroc'] = weighted_auroc
        test_metrics['weighted_auprc'] = weighted_auprc
        test_metrics['per_class_auroc'] = per_class_auroc
        test_metrics['per_class_auprc'] = per_class_auprc

        # # TensorBoard에 Test 메트릭 로깅
        # logger.log_epoch(epoch, 0.0, {
        #     'acc': test_metrics['overall_accuracy'],
        #     'macro_f1': test_metrics['macro_f1'],
        #     'macro_auroc': test_metrics['macro_auroc'],
        #     'macro_auprc': test_metrics['macro_auprc'],
        #     'weighted_auroc': test_metrics['weighted_auroc'],
        #     'weighted_auprc': test_metrics['weighted_auprc']
        # }, phase=f'test_{model_key}')

        # Per-class 메트릭 로깅
        print_per_class_metrics(test_metrics, CLASSES)
        for i, cls in enumerate(CLASSES):
            logger.writer.add_scalar(f'Test_{model_key}/F1_{cls}', test_metrics['per_class_f1'][i], epoch)
            logger.writer.add_scalar(f'Test_{model_key}/AUROC_{cls}', per_class_auroc[i], epoch)
            logger.writer.add_scalar(f'Test_{model_key}/AUPRC_{cls}', per_class_auprc[i], epoch)

        # 결과 저장
        results_dict[model_key] = {
            'best_epoch': epoch,
            'valid_metrics': best[model_key]['valid_metrics'],
            'test_metrics': test_metrics
        }

        print(f"  Test Acc: {test_metrics['overall_accuracy']:.4f}")
        print(f"  Test Macro AUROC: {test_metrics['macro_auroc']:.4f}")
        print(f"  Test Macro AUPRC: {test_metrics['macro_auprc']:.4f}")

    # Best weights 저장
    for model_key in ['auroc', 'auprc', 'last']:
        torch.save({
            'model_state_dict': best[model_key]['state_dict'],
            'epoch': best[model_key]['epoch'],
        }, os.path.join(exp_dir, 'best_weights', f'best_{model_key}_{exp_name}.pth'))

    print(f"\n✅ {exp_name} completed")
    print(f"  Saved: AUROC (epoch {best['auroc']['epoch']}), AUPRC (epoch {best['auprc']['epoch']}), Last (epoch {best['last']['epoch']})")

    return results_dict, exp_dir


def fill_excel_template(all_results, template_path, output_path, append_mode=False):
    """
    엑셀 양식에 결과 채우기

    Args:
        all_results: {exp_name: {model_key: results}} 딕셔너리
        template_path: 양식 파일 경로
        output_path: 출력 파일 경로
        append_mode: True면 기존 파일 열기, False면 템플릿에서 새로 생성
    """
    if append_mode and os.path.exists(output_path):
        wb = load_workbook(output_path)
    else:
        wb = load_workbook(template_path)

    # 시트 가져오기
    ws_perf = wb['Performance Metrics']
    ws_valid = wb['Valid']
    ws_test = wb['Test']
    ws_confusion = wb['Confusion Matrix']

    # ========== Performance Metrics 시트 채우기 ==========
    exp_row_map = {
        'A0@': 10, 'A1@': 11, 'A2@': 12,
        'B0@': 13, 'B1@': 14, 'B2@': 15,
        'A0^': 16, 'A1^': 17, 'A2^': 18,
        'B0^': 19, 'B1^': 20, 'B2^': 21
    }

    for exp_name, model_results in all_results.items():
        # @ 및 ^ 실험에 대해 각각 처리
        for suffix, model_key in [('@', 'auroc'), ('^', 'auprc')]:
            exp_key = exp_name + suffix
            if exp_key not in exp_row_map:
                continue

            metrics = model_results[model_key]['test_metrics']
            row = exp_row_map[exp_key]

            # Macro metrics (col B-F, 2-6)
            ws_perf.cell(row, 2).value = metrics['macro_accuracy']
            ws_perf.cell(row, 3).value = metrics['macro_recall']
            ws_perf.cell(row, 4).value = metrics['macro_specificity']
            ws_perf.cell(row, 5).value = metrics['macro_prec']
            ws_perf.cell(row, 6).value = metrics['macro_f1']

            # Weighted metrics (col G-K, 7-11)
            ws_perf.cell(row, 7).value = metrics['weighted_accuracy']
            ws_perf.cell(row, 8).value = metrics['weighted_recall']
            ws_perf.cell(row, 9).value = metrics['weighted_specificity']
            ws_perf.cell(row, 10).value = metrics['weighted_prec']
            ws_perf.cell(row, 11).value = metrics['weighted_f1']

            # Per-class metrics (N, S, V, F)
            for i, cls in enumerate(CLASSES):
                base_col = 12 + i * 5
                ws_perf.cell(row, base_col).value = metrics['per_class_acc'][i]
                ws_perf.cell(row, base_col + 1).value = metrics['per_class_recall'][i]
                ws_perf.cell(row, base_col + 2).value = metrics['per_class_specificity'][i]
                ws_perf.cell(row, base_col + 3).value = metrics['per_class_precision'][i]
                ws_perf.cell(row, base_col + 4).value = metrics['per_class_f1'][i]

    # ========== Valid 시트 채우기 ==========
    valid_exp_map = {
        'A0@': 3, 'A0^': 4,
        'A1@': 6, 'A1^': 7,
        'A2@': 9, 'A2^': 10,
        'B0@': 12, 'B0^': 13,
        'B1@': 15, 'B1^': 16,
        'B2@': 18, 'B2^': 19,
    }

    for exp_name, model_results in all_results.items():
        for suffix, model_key in [('@', 'auroc'), ('^', 'auprc')]:
            exp_key = exp_name + suffix
            if exp_key not in valid_exp_map:
                continue

            row = valid_exp_map[exp_key]
            valid_metrics = model_results[model_key]['valid_metrics']
            best_epoch = model_results[model_key]['best_epoch']

            # Experiment, Model, Epoch
            ws_valid.cell(row, 1).value = exp_key
            ws_valid.cell(row, 2).value = exp_name  # 모델 타입

            if suffix == '@':
                ws_valid.cell(row, 3).value = f"{best_epoch} (AUROC)"
            else:
                ws_valid.cell(row, 3).value = f"{best_epoch} (AUPRC)"

            # Per-class AUROC (N, S, V, F)
            for i in range(4):
                ws_valid.cell(row, 4+i).value = valid_metrics.get('per_class_auroc', [0,0,0,0])[i]

            # Per-class AUPRC (N, S, V, F)
            for i in range(4):
                ws_valid.cell(row, 8+i).value = valid_metrics.get('per_class_auprc', [0,0,0,0])[i]

            # Overall metrics
            ws_valid.cell(row, 12).value = valid_metrics.get('acc', 0.0)
            ws_valid.cell(row, 13).value = valid_metrics.get('macro_auroc', 0.0)
            ws_valid.cell(row, 14).value = valid_metrics.get('macro_auprc', 0.0)
            ws_valid.cell(row, 15).value = valid_metrics.get('weighted_auroc', 0.0)
            ws_valid.cell(row, 16).value = valid_metrics.get('weighted_auprc', 0.0)

    # ========== Test 시트 채우기 ==========
    for exp_name, model_results in all_results.items():
        for suffix, model_key in [('@', 'auroc'), ('^', 'auprc')]:
            exp_key = exp_name + suffix
            if exp_key not in valid_exp_map:
                continue

            row = valid_exp_map[exp_key]
            test_metrics = model_results[model_key]['test_metrics']
            best_epoch = model_results[model_key]['best_epoch']

            # Experiment, Model, Epoch
            ws_test.cell(row, 1).value = exp_key
            ws_test.cell(row, 2).value = exp_name

            if suffix == '@':
                ws_test.cell(row, 3).value = f"{best_epoch} (AUROC)"
            else:
                ws_test.cell(row, 3).value = f"{best_epoch} (AUPRC)"

            # Per-class AUROC (N, S, V, F)
            for i in range(4):
                ws_test.cell(row, 4+i).value = test_metrics.get('per_class_auroc', [0,0,0,0])[i]

            # Per-class AUPRC (N, S, V, F)
            for i in range(4):
                ws_test.cell(row, 8+i).value = test_metrics.get('per_class_auprc', [0,0,0,0])[i]

            # Overall metrics
            ws_test.cell(row, 12).value = test_metrics.get('overall_accuracy', 0.0)
            ws_test.cell(row, 13).value = test_metrics.get('macro_auroc', 0.0)
            ws_test.cell(row, 14).value = test_metrics.get('macro_auprc', 0.0)
            ws_test.cell(row, 15).value = test_metrics.get('weighted_auroc', 0.0)
            ws_test.cell(row, 16).value = test_metrics.get('weighted_auprc', 0.0)

    # ========== Confusion Matrix 시트 채우기 (3x6 그리드) ==========
    confusion_grid = [
        ['A0@', 'A1@', 'A2@'],
        ['B0@', 'B1@', 'B2@'],
        ['A0^', 'A1^', 'A2^'],
        ['B0^', 'B1^', 'B2^']
    ]

    block_width = 7

    for row_idx, exp_row in enumerate(confusion_grid):
        for col_idx, exp_key in enumerate(exp_row):
            exp_name = exp_key[:-1]  # A0, A1, etc.
            suffix = exp_key[-1]     # @ or ^
            model_key = 'auroc' if suffix == '@' else 'auprc'

            if exp_name not in all_results:
                continue

            test_metrics = all_results[exp_name][model_key]['test_metrics']
            confusion_matrix = test_metrics.get('confusion_matrix', np.zeros((4, 4)))

            block_col = 1 + col_idx * block_width
            base_row = 1 + row_idx * 8 + 3

            # Confusion Matrix 데이터 채우기 (4x4)
            for i in range(4):
                for j in range(4):
                    ws_confusion.cell(base_row + i, block_col + 2 + j).value = int(confusion_matrix[i, j])

    # 저장
    wb.save(output_path)
    print(f"\n✅ Results saved to: {output_path}")


# =============================================================================
# 메인 실행
# =============================================================================

if __name__ == '__main__':
    device = torch.device('cuda' if torch.cuda.is_available() else 'cpu')

    print("="*80)
    print("Automated Experiments (Simplified)")
    print("="*80)
    print(f"Device: {device}")
    if torch.cuda.is_available():
        print(f"GPU: {torch.cuda.get_device_name(0)}")
    print(f"Output: {OUTPUT_PATH}")
    print("="*80)

    # 템플릿 및 출력 경로 설정
    template_path = 'Results_Template.xlsx'
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    output_path = os.path.join(OUTPUT_PATH, f'Results_{timestamp}.xlsx')

    # 실험 실행 및 각 실험마다 저장
    all_results = {}

    for idx, (exp_name, model_type) in enumerate(EXPERIMENTS):
        results_dict, exp_dir = run_experiment(exp_name, model_type, device)
        all_results[exp_name] = results_dict

        # 각 실험 완료 후 즉시 엑셀에 저장
        append_mode = idx > 0  # 첫 번째 실험은 템플릿에서 생성, 이후는 append
        fill_excel_template(all_results, template_path, output_path, append_mode=append_mode)

        print(f"\n💾 Results saved after {exp_name}: {output_path}")
        print(f"   Progress: {idx+1}/{len(EXPERIMENTS)} experiments completed")

    print("\n" + "="*80)
    print("✅ ALL EXPERIMENTS COMPLETED!")
    print("="*80)
    print(f"Final results saved to: {output_path}")
