# main_polyloss_gridsearch.py - Poly Loss Grid Search (자동화 버전)
# 모든 Poly Loss 조합을 자동으로 실험하고 결과를 엑셀에 정리

import os
import time
import copy
from collections import Counter
from datetime import datetime

import numpy as np
import pandas as pd
import torch
from torch.utils.data import DataLoader

from utils import set_seed, load_or_extract_data
from model import get_model
from dataloader import ECGDataset
from train import train_one_epoch, validate, save_model
from test import evaluate, calculate_metrics, print_metrics, save_results_excel, save_confusion_matrix
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from logger import (
    TrainingLogger, print_epoch_header, print_per_class_metrics,
    print_epoch_stats, print_confidence_stats, print_epoch_time
)

# =============================================================================
# 실험 설정
# =============================================================================

# 공통 설정
DATA_PATH = './data/mit-bih-arrhythmia-database-1.0.0/'
OUTPUT_PATH = './gridsearch_results_tfstyle/'
BATCH_SIZE = 1024
EPOCHS = 50
LR = 0.0001
WEIGHT_DECAY = 1e-3
SEED = 1234
CLASSES = ['N', 'S', 'V', 'F']


# 모델 설정 (config.py의 EXP_NAME에 따라 모델 선택)
BASE_MODEL = 'cross_attention'  # baseline, naive_concatenate, cross_attention, *_B
RR_FEATURE_OPTION = "opt3"
RR_FEATURE_DIMS = {"opt1": 7, "opt2": 38, "opt3": 7, "opt4": 7}

MODEL_CONFIG = {
    'in_channels': 1,
    'out_ch': 180,
    'mid_ch': 30,
    'num_heads': 9,
    'n_rr': RR_FEATURE_DIMS[RR_FEATURE_OPTION],
}

# ECG Parameters
VALID_LEADS = ['MLII', 'V1', 'V2', 'V4', 'V5']
OUT_LEN = 720

# 데이터 분할
DS1_TRAIN = [
    '101', '106', '108', '109', '112', '115', '116', '118', '119',
    '122', '201', '203', '209', '215', '223', '230', '208'
]
DS1_VALID = ['114', '124', '205', '207', '220']
DS2_TEST = [
    '100', '103', '105', '111', '113', '117', '121', '123', '200', '202',
    '210', '212', '213', '214', '219', '221', '222', '228', '231', '232',
    '233', '234'
]

# =============================================================================
# ★ Grid Search 설정
# =============================================================================

def create_phase1_experiments():
    """Phase 1: Baseline + Poly-1 Grid Search"""
    experiments = []

    # Baseline: Cross-Entropy (alpha1=0, alpha2=0)
    experiments.append({
        'alpha1': 0.0,
        'alpha2': 0.0,
        'name': 'CE_baseline',
        'description': 'Cross-Entropy Loss (Baseline)',
        'phase': 1
    })

    # Focal: (alpha1=-1, alpha2=0)
    experiments.append({
        'alpha1': -1.0,
        'alpha2': 0.0,
        'name': 'Focal',
        'description': '(Drop Linear)',
        'phase': 1
    })

    # Poly-1 Grid Search
    for a1 in [-2.0, -1.8, -1.6, -1.4, -1.2, -1.0, -0.8, -0.6, -0.4, -0.2, 0.2, 0.4, 0.6, 0.8, 1.0, 1.2, 1.4, 1.6, 1.8, 2.0]:
        experiments.append({
            'alpha1': a1,
            'alpha2': 0.0,
            'name': f'Poly1_a{a1:.1f}',
            'description': f'Poly-1 Loss (α₁={a1})',
            'phase': 1
        })

    return experiments


def create_phase2_experiments(best_alpha1):
    """Phase 2: Poly-2 Grid Search with best alpha1 from Phase 1"""
    experiments = []

    # Poly-2 Grid Search
    for a2 in [-2.0, -1.8, -1.6, -1.4, -1.2, -1.0, -0.8, -0.6, -0.4, -0.2, 0.0, 0.2, 0.4, 0.6, 0.8, 1.0, 1.2, 1.4, 1.6, 1.8, 2.0]:
        experiments.append({
            'alpha1': best_alpha1,
            'alpha2': a2,
            'name': f'Poly2_a{a2:.1f}',
            'description': f'Poly-2 Loss (α₁={best_alpha1}, α₂={a2})',
            'phase': 2
        })

    return experiments

# =============================================================================
# 유틸리티 함수
# =============================================================================

def create_grid_exp_dir(exp_name):
    """Grid Search 실험용 폴더 생성"""
    os.makedirs(OUTPUT_PATH, exist_ok=True)
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    exp_dir = os.path.join(OUTPUT_PATH, f'{exp_name}_{timestamp}')
    os.makedirs(exp_dir, exist_ok=True)
    os.makedirs(os.path.join(exp_dir, 'best_weights'), exist_ok=True)
    return exp_dir


def run_single_experiment(exp_config, train_loader, valid_loader, test_loader, 
                          n_records, device):
    """
    단일 Grid Search 실험 수행
    
    Args:
        exp_config: 실험 설정 딕셔너리
        train_loader, valid_loader, test_loader: DataLoader들
        n_records: 전체 환자 수
        device: torch device
    
    Returns:
        results: 실험 결과 딕셔너리
    """
    ALPHA1 = exp_config['alpha1']
    ALPHA2 = exp_config['alpha2']
    exp_name = exp_config['name']
    
    print(f"\n{'='*80}")
    print(f"Experiment: {exp_name}")
    print(f"Description: {exp_config['description']}")
    print(f"Parameters: α₁={ALPHA1}, α₂={ALPHA2}")
    print(f"{'='*80}")
    
    set_seed(SEED)
    exp_dir = create_grid_exp_dir(exp_name)
    
    # TensorBoard Logger
    logger = TrainingLogger(os.path.join(exp_dir, 'runs'))
    
    # 모델 생성
    model = get_model(
        exp_name=BASE_MODEL,
        nOUT=len(CLASSES),
        n_pid=n_records,
        **MODEL_CONFIG
    ).to(device)
    
    # 학습 설정
    optimizer = torch.optim.AdamW(model.parameters(), lr=LR, weight_decay=WEIGHT_DECAY)
    scheduler = torch.optim.lr_scheduler.StepLR(optimizer, step_size=20, gamma=0.5)
    
    # Best 모델 추적
    best = {
        'auroc': {'value': 0.0, 'epoch': 0, 'state_dict': None, 'valid_metrics': None},
        'auprc': {'value': 0.0, 'epoch': 0, 'state_dict': None, 'valid_metrics': None},
    }
    
    # 학습 루프
    print_epoch_header()
    exp_start = time.time()
    
    for epoch in range(1, EPOCHS + 1):
        epoch_start = time.time()
        
        # Train
        train_loss, train_metrics, p_t_n_train, p_t_s_train, p_t_v_train, p_t_f_train = train_one_epoch(
            model, train_loader, ALPHA1, ALPHA2, optimizer, device
        )
        current_lr = optimizer.param_groups[0]['lr']

        if epoch % 10 == 0 or epoch == 1:
            print_epoch_stats(epoch, train_loss, train_metrics['acc'], current_lr, phase='Train')
            print_per_class_metrics(train_metrics, CLASSES, phase='Train')
            print_confidence_stats(p_t_n_train, p_t_s_train, p_t_v_train, p_t_f_train, phase='Train')

        logger.log_epoch(epoch, train_loss, train_metrics, phase='train')
        logger.log_confidence(epoch, p_t_n_train, p_t_s_train, p_t_v_train, p_t_f_train, phase='train')

        # Validation
        valid_loss, valid_metrics, p_t_n_valid, p_t_s_valid, p_t_v_valid, p_t_f_valid = validate(
            model, valid_loader, ALPHA1, ALPHA2, device
        )

        if epoch % 10 == 0 or epoch == 1:
            print_epoch_stats(epoch, valid_loss, valid_metrics['acc'], current_lr, phase='Valid')
            print_per_class_metrics(valid_metrics, CLASSES, phase='Valid')
            print_confidence_stats(p_t_n_valid, p_t_s_valid, p_t_v_valid, p_t_f_valid, phase='Valid')

        logger.log_epoch(epoch, valid_loss, valid_metrics, phase='valid')
        logger.log_confidence(epoch, p_t_n_valid, p_t_s_valid, p_t_v_valid, p_t_f_valid, phase='valid')
        
        # Best AUROC 체크
        if valid_metrics['macro_auroc'] > best['auroc']['value']:
            best['auroc'] = {
                'value': valid_metrics['macro_auroc'],
                'epoch': epoch,
                'state_dict': copy.deepcopy(model.state_dict()),
                'valid_metrics': copy.deepcopy(valid_metrics)
            }
            if epoch % 10 == 0 or epoch == 1:
                print(f"  ★ [BEST AUROC] {best['auroc']['value']:.4f}")
        
        # Best AUPRC 체크
        if valid_metrics['macro_auprc'] > best['auprc']['value']:
            best['auprc'] = {
                'value': valid_metrics['macro_auprc'],
                'epoch': epoch,
                'state_dict': copy.deepcopy(model.state_dict()),
                'valid_metrics': copy.deepcopy(valid_metrics)
            }
            if epoch % 10 == 0 or epoch == 1:
                print(f"  ★ [BEST AUPRC] {best['auprc']['value']:.4f}")
        
        scheduler.step()
        
        if epoch % 10 == 0 or epoch == 1:
            print_epoch_time(epoch, time.time() - epoch_start)
            print("=" * 80)
    
    logger.close()
    
    # AUROC 기준 모델 테스트
    print(f"\n--- Testing AUROC Best Model (Epoch {best['auroc']['epoch']}) ---")
    model.load_state_dict(best['auroc']['state_dict'])
    model.eval()

    y_pred, y_true, eval_results = evaluate(model, test_loader, device)
    test_metrics_auroc = calculate_metrics(np.array(y_true), np.array(y_pred))

    # AUROC/AUPRC 계산을 위한 확률값 추출
    y_probs_auroc = np.array([item['all_probs'] for item in eval_results])

    from logger import calculate_auprc, calculate_auroc
    num_classes = len(CLASSES)
    macro_auprc_auroc, weighted_auprc_auroc, per_class_auprc_auroc = calculate_auprc(
        np.array(y_true), y_probs_auroc, num_classes
    )
    macro_auroc_auroc, weighted_auroc_auroc, per_class_auroc_auroc = calculate_auroc(
        np.array(y_true), y_probs_auroc, num_classes
    )

    # 메트릭에 추가
    test_metrics_auroc['macro_auroc'] = macro_auroc_auroc
    test_metrics_auroc['macro_auprc'] = macro_auprc_auroc
    test_metrics_auroc['weighted_auroc'] = weighted_auroc_auroc
    test_metrics_auroc['weighted_auprc'] = weighted_auprc_auroc
    test_metrics_auroc['per_class_auroc'] = per_class_auroc_auroc
    test_metrics_auroc['per_class_auprc'] = per_class_auprc_auroc

    # AUPRC 기준 모델 테스트
    print(f"\n--- Testing AUPRC Best Model (Epoch {best['auprc']['epoch']}) ---")
    model.load_state_dict(best['auprc']['state_dict'])
    model.eval()

    y_pred, y_true, eval_results = evaluate(model, test_loader, device)
    test_metrics_auprc = calculate_metrics(np.array(y_true), np.array(y_pred))

    # AUROC/AUPRC 계산을 위한 확률값 추출
    y_probs_auprc = np.array([item['all_probs'] for item in eval_results])

    macro_auprc_auprc, weighted_auprc_auprc, per_class_auprc_auprc = calculate_auprc(
        np.array(y_true), y_probs_auprc, num_classes
    )
    macro_auroc_auprc, weighted_auroc_auprc, per_class_auroc_auprc = calculate_auroc(
        np.array(y_true), y_probs_auprc, num_classes
    )

    # 메트릭에 추가
    test_metrics_auprc['macro_auroc'] = macro_auroc_auprc
    test_metrics_auprc['macro_auprc'] = macro_auprc_auprc
    test_metrics_auprc['weighted_auroc'] = weighted_auroc_auprc
    test_metrics_auprc['weighted_auprc'] = weighted_auprc_auprc
    test_metrics_auprc['per_class_auroc'] = per_class_auroc_auprc
    test_metrics_auprc['per_class_auprc'] = per_class_auprc_auprc
    
    # Test 결과 TensorBoard 로깅
    logger = TrainingLogger(os.path.join(exp_dir, 'runs'))
    logger.log_epoch(best['auroc']['epoch'], 0.0, test_metrics_auroc, phase='test_auroc')
    logger.log_epoch(best['auprc']['epoch'], 0.0, test_metrics_auprc, phase='test_auprc')
    logger.close()
    
    # 결과 저장
    save_results_excel(test_metrics_auprc, CLASSES,
                       os.path.join(exp_dir, f'results_auprc_{exp_name}.xlsx'))
    save_confusion_matrix(test_metrics_auprc['confusion_matrix'], CLASSES,
                          os.path.join(exp_dir, f'confusion_matrix_auprc_{exp_name}.png'))

    save_results_excel(test_metrics_auroc, CLASSES,
                       os.path.join(exp_dir, f'results_auroc_{exp_name}.xlsx'))
    save_confusion_matrix(test_metrics_auroc['confusion_matrix'], CLASSES,
                          os.path.join(exp_dir, f'confusion_matrix_auroc_{exp_name}.png'))

    # Confusion Matrix 저장 (numpy array로)
    test_metrics_auprc['confusion_matrix_array'] = test_metrics_auprc['confusion_matrix']
    test_metrics_auroc['confusion_matrix_array'] = test_metrics_auroc['confusion_matrix']
    
    # Best weights 저장
    torch.save({
        'model_state_dict': best['auroc']['state_dict'],
        'epoch': best['auroc']['epoch'],
        'auroc': best['auroc']['value']
    }, os.path.join(exp_dir, 'best_weights', f'best_auroc_{exp_name}.pth'))
    
    torch.save({
        'model_state_dict': best['auprc']['state_dict'],
        'epoch': best['auprc']['epoch'],
        'auprc': best['auprc']['value']
    }, os.path.join(exp_dir, 'best_weights', f'best_auprc_{exp_name}.pth'))
    
    # 결과 반환
    results = {
        'exp_name': exp_name,
        'description': exp_config['description'],
        'alpha1': ALPHA1,
        'alpha2': ALPHA2,
        'phase': exp_config['phase'],  # Phase 정보 추가

        # Valid - AUROC Best
        'valid_auroc_epoch': best['auroc']['epoch'],
        'valid_auroc': best['auroc']['value'],
        'valid_auroc_auprc': best['auroc']['valid_metrics']['macro_auprc'],
        'valid_auroc_f1': best['auroc']['valid_metrics']['macro_f1'],
        'valid_auroc_acc': best['auroc']['valid_metrics']['acc'],

        # Valid - AUPRC Best
        'valid_auprc_epoch': best['auprc']['epoch'],
        'valid_auprc': best['auprc']['value'],
        'valid_auprc_auroc': best['auprc']['valid_metrics']['macro_auroc'],
        'valid_auprc_f1': best['auprc']['valid_metrics']['macro_f1'],
        'valid_auprc_acc': best['auprc']['valid_metrics']['acc'],

        # Test - AUROC Best Model
        'test_auroc_acc': test_metrics_auroc['overall_accuracy'],
        'test_auroc_macro_f1': test_metrics_auroc['macro_f1'],
        'test_auroc_macro_auroc': test_metrics_auroc['macro_auroc'],
        'test_auroc_macro_auprc': test_metrics_auroc['macro_auprc'],
        'test_auroc_macro_precision': test_metrics_auroc['macro_precision'],
        'test_auroc_macro_recall': test_metrics_auroc['macro_recall'],
        'test_auroc_macro_specificity': test_metrics_auroc['macro_specificity'],
        'test_auroc_weighted_f1': test_metrics_auroc['weighted_f1'],
        'test_auroc_weighted_acc': test_metrics_auroc['weighted_accuracy'],
        'test_auroc_weighted_precision': test_metrics_auroc['weighted_precision'],
        'test_auroc_weighted_recall': test_metrics_auroc['weighted_recall'],
        'test_auroc_weighted_specificity': test_metrics_auroc['weighted_specificity'],
        'test_auroc_confusion': test_metrics_auroc['confusion_matrix_array'],

        # Test - AUPRC Best Model
        'test_auprc_acc': test_metrics_auprc['overall_accuracy'],
        'test_auprc_macro_f1': test_metrics_auprc['macro_f1'],
        'test_auprc_macro_auroc': test_metrics_auprc['macro_auroc'],
        'test_auprc_macro_auprc': test_metrics_auprc['macro_auprc'],
        'test_auprc_macro_precision': test_metrics_auprc['macro_precision'],
        'test_auprc_macro_recall': test_metrics_auprc['macro_recall'],
        'test_auprc_macro_specificity': test_metrics_auprc['macro_specificity'],
        'test_auprc_weighted_f1': test_metrics_auprc['weighted_f1'],
        'test_auprc_weighted_acc': test_metrics_auprc['weighted_accuracy'],
        'test_auprc_weighted_precision': test_metrics_auprc['weighted_precision'],
        'test_auprc_weighted_recall': test_metrics_auprc['weighted_recall'],
        'test_auprc_weighted_specificity': test_metrics_auprc['weighted_specificity'],
        'test_auprc_confusion': test_metrics_auprc['confusion_matrix_array'],

        # Per-class Test (AUPRC Best Model)
        'test_N_f1': test_metrics_auprc['per_class_f1'][0],
        'test_S_f1': test_metrics_auprc['per_class_f1'][1],
        'test_V_f1': test_metrics_auprc['per_class_f1'][2],
        'test_F_f1': test_metrics_auprc['per_class_f1'][3],

        'time_min': (time.time() - exp_start) / 60,
        'exp_dir': exp_dir
    }
    
    print(f"\n✅ {exp_name} completed in {results['time_min']:.1f} min")
    print(f"  Valid AUROC: {best['auroc']['value']:.4f} (epoch {best['auroc']['epoch']})")
    print(f"  Valid AUPRC: {best['auprc']['value']:.4f} (epoch {best['auprc']['epoch']})")
    print(f"  Test Acc (AUPRC model): {test_metrics_auprc['overall_accuracy']:.4f}")
    
    return results


def create_summary_excel(all_results, output_path):
    """
    Grid Search 결과 요약 엑셀 생성 (Confusion Matrix 포함)

    Args:
        all_results: 모든 실험 결과 리스트
        output_path: 출력 파일 경로
    """
    # DataFrame 생성 (confusion matrix 제외)
    df_data = []
    for r in all_results:
        r_copy = r.copy()
        # Confusion matrix는 별도 시트로 처리
        r_copy.pop('test_auroc_confusion', None)
        r_copy.pop('test_auprc_confusion', None)
        df_data.append(r_copy)

    df = pd.DataFrame(df_data)

    # 정렬 (Valid AUPRC 기준)
    df_sorted = df.sort_values('valid_auprc', ascending=False).reset_index(drop=True)
    df_sorted.insert(0, 'Rank', range(1, len(df_sorted) + 1))

    # 엑셀 저장 (여러 시트)
    with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
        # Sheet 1: 전체 요약 (Valid AUPRC 순)
        summary_cols = [
            'Rank', 'exp_name', 'description', 'alpha1', 'alpha2',
            'valid_auprc_epoch', 'valid_auprc', 'valid_auprc_auroc', 'valid_auprc_f1',
            'test_auprc_acc', 'test_auprc_macro_f1', 'test_auprc_macro_auroc', 'test_auprc_macro_auprc',
            'time_min'
        ]
        df_sorted[summary_cols].to_excel(writer, sheet_name='Summary_by_AUPRC', index=False)

        # Sheet 2: Valid AUROC 순
        df_auroc = df.sort_values('valid_auroc', ascending=False).reset_index(drop=True)
        df_auroc.insert(0, 'Rank', range(1, len(df_auroc) + 1))
        auroc_cols = [
            'Rank', 'exp_name', 'description', 'alpha1', 'alpha2',
            'valid_auroc_epoch', 'valid_auroc', 'valid_auroc_auprc', 'valid_auroc_f1',
            'test_auroc_acc', 'test_auroc_macro_f1', 'test_auroc_macro_auroc', 'test_auroc_macro_auprc',
            'time_min'
        ]
        df_auroc[auroc_cols].to_excel(writer, sheet_name='Summary_by_AUROC', index=False)

        # Sheet 3: Test Performance Comparison
        test_cols = [
            'exp_name', 'alpha1', 'alpha2',
            'test_auprc_acc', 'test_auprc_macro_f1', 'test_auprc_macro_auroc', 'test_auprc_macro_auprc',
            'test_N_f1', 'test_S_f1', 'test_V_f1', 'test_F_f1'
        ]
        df_sorted[test_cols].to_excel(writer, sheet_name='Test_Performance', index=False)

        # Sheet 4: Complete Results
        df_sorted.to_excel(writer, sheet_name='All_Results', index=False)

        # Sheet 5: Confusion Matrices (AUPRC Best)
        ws_confusion = writer.book.create_sheet('Confusion_Matrices_AUPRC')
        create_confusion_sheet(ws_confusion, all_results, 'auprc')

        # Sheet 6: Confusion Matrices (AUROC Best)
        ws_confusion_auroc = writer.book.create_sheet('Confusion_Matrices_AUROC')
        create_confusion_sheet(ws_confusion_auroc, all_results, 'auroc')

    print(f"\n📊 Summary Excel saved: {output_path}")


def create_confusion_sheet(ws, all_results, metric_type='auprc'):
    """
    Confusion Matrix 시트 생성 - 가로 배치 (3열 그리드)

    Args:
        ws: worksheet
        all_results: 모든 실험 결과 리스트
        metric_type: 'auprc' or 'auroc'
    """
    # 헤더 스타일
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    subheader_fill = PatternFill(start_color="B7DEE8", end_color="B7DEE8", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    subheader_font = Font(bold=True, size=10)
    center_align = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    classes = ['N', 'S', 'V', 'F']
    block_width = 7  # 각 블록의 너비
    blocks_per_row = 3  # 한 행에 3개 블록

    current_row = 1
    col_idx = 0

    for result in all_results:
        exp_name = result['exp_name']
        confusion_key = f'test_{metric_type}_confusion'

        if confusion_key not in result:
            continue

        confusion_matrix = result[confusion_key]

        # 현재 블록의 시작 열
        block_col = 1 + col_idx * block_width

        # 실험명 헤더
        ws.merge_cells(start_row=current_row, start_column=block_col,
                      end_row=current_row, end_column=block_col+5)
        cell = ws.cell(row=current_row, column=block_col)
        cell.value = f'{exp_name} ({metric_type.upper()})'
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_align
        cell.border = thin_border

        # "Predicted" 라벨
        ws.merge_cells(start_row=current_row+1, start_column=block_col+2,
                      end_row=current_row+1, end_column=block_col+5)
        cell = ws.cell(row=current_row+1, column=block_col+2)
        cell.value = 'Predicted'
        cell.fill = subheader_fill
        cell.font = subheader_font
        cell.alignment = center_align
        cell.border = thin_border

        # 클래스 헤더 (Predicted)
        for i, cls in enumerate(classes):
            cell = ws.cell(row=current_row+2, column=block_col+2+i)
            cell.value = cls
            cell.fill = subheader_fill
            cell.font = subheader_font
            cell.alignment = center_align
            cell.border = thin_border

        # "Actual" 라벨 (병합)
        ws.merge_cells(start_row=current_row+3, start_column=block_col,
                      end_row=current_row+6, end_column=block_col)
        cell = ws.cell(row=current_row+3, column=block_col)
        cell.value = 'Actual'
        cell.fill = subheader_fill
        cell.font = subheader_font
        cell.alignment = center_align
        cell.border = thin_border

        # Confusion Matrix 데이터
        for i, cls in enumerate(classes):
            # 클래스 이름
            cell = ws.cell(row=current_row+3+i, column=block_col+1)
            cell.value = cls
            cell.fill = PatternFill(start_color="DCE6F1", end_color="DCE6F1", fill_type="solid")
            cell.font = subheader_font
            cell.alignment = center_align
            cell.border = thin_border

            # 데이터 셀 (정수)
            for j in range(4):
                cell = ws.cell(row=current_row+3+i, column=block_col+2+j)
                cell.value = int(confusion_matrix[i, j])
                cell.number_format = '0'
                cell.alignment = center_align
                cell.border = thin_border

        # 다음 블록 위치 계산
        col_idx += 1
        if col_idx >= blocks_per_row:
            col_idx = 0
            current_row += 8  # 다음 행으로 이동

    # 열 너비 조정
    for i in range(blocks_per_row):
        block_col = 1 + i * block_width
        ws.column_dimensions[get_column_letter(block_col)].width = 8  # Actual
        ws.column_dimensions[get_column_letter(block_col+1)].width = 6  # Class
        for j in range(4):  # Data columns
            ws.column_dimensions[get_column_letter(block_col+2+j)].width = 9


# =============================================================================
# 메인 실행
# =============================================================================

if __name__ == '__main__':
    device = torch.device('cuda' if torch.cuda.is_available() else 'cpu')

    print("="*80)
    print("Poly Loss Grid Search - Two-Phase Automated")
    print("="*80)
    print(f"Device: {device}")
    if torch.cuda.is_available():
        print(f"GPU: {torch.cuda.get_device_name(0)}")
    print(f"Base Model: {BASE_MODEL}")
    print(f"Output: {OUTPUT_PATH}")
    print("="*80)
    
    # 데이터 로드 (한번만)
    print("\n[1/2] Loading data...")
    
    train_data, train_labels, train_rr, train_pid, train_sid = load_or_extract_data(
        record_list=DS1_TRAIN, base_path=DATA_PATH, valid_leads=VALID_LEADS,
        out_len=OUT_LEN, split_name="Train"
    )
    valid_data, valid_labels, valid_rr, valid_pid, valid_sid = load_or_extract_data(
        record_list=DS1_VALID, base_path=DATA_PATH, valid_leads=VALID_LEADS,
        out_len=OUT_LEN, split_name="Valid"
    )
    test_data, test_labels, test_rr, test_pid, test_sid = load_or_extract_data(
        record_list=DS2_TEST, base_path=DATA_PATH, valid_leads=VALID_LEADS,
        out_len=OUT_LEN, split_name="Test"
    )
    
    # DataLoader 생성
    train_dataset = ECGDataset(train_data, train_rr, train_labels, train_pid, train_sid)
    valid_dataset = ECGDataset(valid_data, valid_rr, valid_labels, valid_pid, valid_sid)
    test_dataset = ECGDataset(test_data, test_rr, test_labels, test_pid, test_sid)
    
    def worker_init_fn(worker_id):
        np.random.seed(SEED + worker_id)
    
    train_loader = DataLoader(train_dataset, batch_size=BATCH_SIZE, shuffle=True,
                              num_workers=4, pin_memory=True, worker_init_fn=worker_init_fn)
    valid_loader = DataLoader(valid_dataset, batch_size=BATCH_SIZE, shuffle=False,
                              num_workers=4, pin_memory=True, worker_init_fn=worker_init_fn)
    test_loader = DataLoader(test_dataset, batch_size=BATCH_SIZE, shuffle=False,
                             num_workers=4, pin_memory=True, worker_init_fn=worker_init_fn)
    
    print(f"  Train: {len(train_labels):,} samples | {dict(Counter(train_labels))}")
    print(f"  Valid: {len(valid_labels):,} samples | {dict(Counter(valid_labels))}")
    print(f"  Test : {len(test_labels):,} samples | {dict(Counter(test_labels))}")
    
    n_records = len(DS1_TRAIN) + len(DS1_VALID)

    # ========================================================================
    # Phase 1: Baseline + Poly-1 Grid Search
    # ========================================================================
    print("\n" + "="*80)
    print("PHASE 1: Baseline + Poly-1 Grid Search")
    print("="*80)

    phase1_experiments = create_phase1_experiments()
    print(f"[2/3] Running {len(phase1_experiments)} Phase 1 experiments...")

    all_results = []
    phase1_start = time.time()

    for exp_idx, exp_config in enumerate(phase1_experiments, 1):
        print(f"\n{'#'*80}")
        print(f"# Phase 1 Progress: [{exp_idx}/{len(phase1_experiments)}]")
        print(f"{'#'*80}")

        try:
            results = run_single_experiment(
                exp_config, train_loader, valid_loader, test_loader,
                n_records, device
            )
            all_results.append(results)
        except Exception as e:
            print(f"\n❌ Error in {exp_config['name']}: {e}")
            import traceback
            traceback.print_exc()
            continue

    phase1_time = (time.time() - phase1_start) / 60

    # Phase 1 결과에서 최고 성능 alpha1 찾기
    print("\n" + "="*80)
    print("PHASE 1 COMPLETED - Finding Best Alpha1")
    print("="*80)

    # ==============Valid AUROC 기준으로 최고 alpha1 선택==========================
    best_alpha1 = 0.0
    best_auroc = 0.0
    best_exp_name = ""

    for result in all_results:
        if result['phase'] == 1 and result['valid_auroc'] > best_auroc:
            best_auroc = result['valid_auroc']
            best_alpha1 = result['alpha1']
            best_exp_name = result['exp_name']

    print(f"✅ Best Alpha1: {best_alpha1:.1f}")
    print(f"   Experiment: {best_exp_name}")
    print(f"   Valid AUROC: {best_auroc:.4f}")
    print(f"   Phase 1 Time: {phase1_time:.1f} min")

    # ========================================================================
    # Phase 2: Poly-2 Grid Search with Best Alpha1
    # ========================================================================
    print("\n" + "="*80)
    print(f"PHASE 2: Poly-2 Grid Search (Alpha1={best_alpha1:.1f})")
    print("="*80)

    phase2_experiments = create_phase2_experiments(best_alpha1)
    print(f"[3/3] Running {len(phase2_experiments)} Phase 2 experiments...")

    phase2_start = time.time()

    for exp_idx, exp_config in enumerate(phase2_experiments, 1):
        print(f"\n{'#'*80}")
        print(f"# Phase 2 Progress: [{exp_idx}/{len(phase2_experiments)}]")
        print(f"{'#'*80}")

        try:
            results = run_single_experiment(
                exp_config, train_loader, valid_loader, test_loader,
                n_records, device
            )
            all_results.append(results)
        except Exception as e:
            print(f"\n❌ Error in {exp_config['name']}: {e}")
            import traceback
            traceback.print_exc()
            continue

    phase2_time = (time.time() - phase2_start) / 60
    total_time = phase1_time + phase2_time
    
    # 결과 요약 엑셀 생성
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    summary_path = os.path.join(OUTPUT_PATH, f'GridSearch_Summary_{timestamp}.xlsx')
    create_summary_excel(all_results, summary_path)
    
    # 최종 요약 출력
    print("\n" + "="*80)
    print("TWO-PHASE GRID SEARCH COMPLETED!")
    print("="*80)
    print(f"Phase 1 time: {phase1_time:.1f} min ({phase1_time/60:.1f} hours)")
    print(f"Phase 2 time: {phase2_time:.1f} min ({phase2_time/60:.1f} hours)")
    print(f"Total time: {total_time:.1f} min ({total_time/60:.1f} hours)")
    print(f"Completed experiments: {len(all_results)} (Phase 1: {len(phase1_experiments)}, Phase 2: {len(phase2_experiments)})")
    print(f"Best Alpha1 used in Phase 2: {best_alpha1:.1f}")
    
    # Top 5 결과 출력
    if all_results:
        sorted_results = sorted(all_results, key=lambda x: x['valid_auprc'], reverse=True)
        
        print(f"\n🏆 Top 5 Results (by Valid AUPRC):")
        print(f"{'Rank':<6} {'Experiment':<25} {'α₁':<8} {'α₂':<8} {'V.AUPRC':<10} {'T.Acc':<10} {'T.F1':<10}")
        print("-" * 90)
        
        for i, r in enumerate(sorted_results[:5], 1):
            print(f"{i:<6} {r['exp_name']:<25} {r['alpha1']:<8.1f} {r['alpha2']:<8.1f} "
                  f"{r['valid_auprc']:<10.4f} {r['test_auprc_acc']:<10.4f} {r['test_auprc_macro_f1']:<10.4f}")
        
        print(f"\n📈 Best Valid AUPRC: {sorted_results[0]['valid_auprc']:.4f} ({sorted_results[0]['exp_name']})")
        print(f"📈 Best Test Acc: {max(all_results, key=lambda x: x['test_auprc_acc'])['test_auprc_acc']:.4f}")
        print(f"📈 Best Test F1: {max(all_results, key=lambda x: x['test_auprc_macro_f1'])['test_auprc_macro_f1']:.4f}")
    
    print(f"\n📁 All results saved to: {OUTPUT_PATH}")
    print(f"📊 Summary: {summary_path}")
    print("="*80)
